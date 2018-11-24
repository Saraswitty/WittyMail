#!/usr/bin/env python
# coding=utf-8

import openpyxl
import emailapi
import re
import util.logger as logger

log = logger.get_logger(__name__)

# Names of the fodder to be used to create the emails
email_fodder_names = []
# Fodder content used to create the emails
email_fodder = []

# Index of the emailid in the email_fodder_names[]. The emailid will be used in the 'to' field
EMAIL_FODDER_TO_INDEX = 0 
# Index of the emailid in the email_fodder_names[]. The emailid will be used in the 'cc' field
EMAIL_FODDER_CC_INDEX = 0

# Extended fodder names that will be appended to the fodder names provided by the user (i.e. email_fodder_names)
extended_email_fodder_names = ['email_subject', 'email_body', 'email_attachment', 'email_sent']
extended_email_fodder_names_EMAIL_BODY_INDEX = -3
extended_email_fodder_names_EMAIL_SUBJECT_INDEX = -4
extended_email_fodder_names_EMAIL_ATTACHMENT_INDEX = -2

# Extended fodder that will be appended to the fodder provided by the user (i.e. email_fodder[])
extended_default_email_fodder = [None, None, None, False]

# TODO Add check to detect extention type
def save_fodder_from_file(loc, email_fodder_names_template = None):
    global email_fodder
    global email_fodder_names
   
    _email_fodder = []  

    # TODO Check if we should read wb_obj.active or wb_obj[0]
    wb_obj = openpyxl.load_workbook(loc)   
    sheet_obj = wb_obj.active

    for i in range(1, sheet_obj.max_row + 1):
        cells = []

        for j in range(1, sheet_obj.max_column + 1):
            cells.append(str(sheet_obj.cell(row = i, column = j).value).encode('UTF8'))
        _email_fodder.append(cells)

    assert email_fodder_names_template == None or email_fodder_names == email_fodder_names_template, \
           "**** The headers in the excel sheet seems to be invalid ****"
    assert len(_email_fodder) > 1,            \
           "**** There are no entries in the excel sheet! ****"

    email_fodder_names = _email_fodder[0]
    email_fodder_names.extend(extended_email_fodder_names)

    email_fodder = _email_fodder[1:]

    for r in email_fodder:
        extended_email_fodder = list(extended_default_email_fodder)
        # TODO Add template support for attachment and support different filetypes
        _attachments = r[1].split(',')
        attachments = [a + ".pdf" for a in _attachments]
        extended_email_fodder[2] = r[1] + b".pdf"
        r.extend(extended_email_fodder)

def get_email_fodder_names():
    return email_fodder_names

def get_email_fodder():
    return email_fodder

def _template_to_str(_st, l):
  log.debug('_template_to_str() str = %s' % (_st))
  st = ''.join(_st)

  str_to_replace = list(set(re.findall(r'#\d+', st)))

  def remove_hash(s):
    return s[1:]

  index = map(int, [remove_hash(s) for s in str_to_replace])

  for i in range(len(index)):
    st = st.replace(str_to_replace[i], l[index[i]])

  log.debug('_template_to_str() final str = %s' % st)
  return st

def save_extended_fodder(to_index, cc_index, subject_template, body_template):
  global EMAIL_FODDER_TO_INDEX
  global EMAIL_FODDER_CC_INDEX

  EMAIL_FODDER_TO_INDEX = to_index
  EMAIL_FODDER_CC_INDEX = cc_index

  for r in email_fodder:
    r[extended_email_fodder_names_EMAIL_SUBJECT_INDEX] = _template_to_str(subject_template, r)
    r[extended_email_fodder_names_EMAIL_BODY_INDEX] = _template_to_str(body_template, r)

def save_attachment_dir(dir_loc, filename_mapping = None):
    global attachment_dir 
    attachment_dir = dir_loc

def send_email(tos = None):
  for e in email_fodder:
    if not tos:
      tos = email_fodder[EMAIL_FODDER_TO_INDEX]
      ccs = email_fodder[EMAIL_FODDER_CC_INDEX]
    else:
      log.debug('Test mail to be sent to+cc %s' % (''.join(tos)))
      ccs = tos

    email_subject = e[extended_email_fodder_names_EMAIL_SUBJECT_INDEX]
    email_body = e[extended_email_fodder_names_EMAIL_BODY_INDEX]
    
    attachments = attachment_dir + a for a in e[extended_email_fodder_names_EMAIL_ATTACHMENT_INDEX]]

    return emailapi.send_email("ajaynair59@gmail.com", tos, email_subject, email_body, ccs, attachments)

def set_login_details(username, password, server="smtp.gmail.com", port=587):
  assert username is not None and password is not None,\
  log.error('server, port number, username or password is None')

  return emailapi.set_login_details(server, port, username, password)
