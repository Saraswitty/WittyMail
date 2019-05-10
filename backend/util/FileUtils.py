import util.logger as logger
from enum import Enum
import PyPDF2
import os
import shutil
from fuzzywuzzy import fuzz
import operator
import openpyxl
import re
import pdb
import math

log = logger.get_logger(__name__)

'''
Used to denote the direction of the rotation of pdf
'''
class Direction(Enum):
    CLOCKWISE = 270 # Degrees to rotate
    ANTICLOCKWISE = 90

    @classmethod
    def convert_from_string(cls, direction_str):
        mapping = {
            'CLOCKWISE': Direction.CLOCKWISE,
            'ANTICLOCKWISE': Direction.ANTICLOCKWISE
        }

        return mapping[direction_str]

'''
All file related helper operations
'''
class FileUtils:
    def __init__(self):
        pass
    
    '''
    Rotate a pdf based on the direction provided
    '''
    def pdf_rotate(self, filepath, direction):
        """
        Rotate all pages in the PDF and replace file on disk

        :param filepath: Full path to PDF file
        :param direction: Direction enum value
        """
        log.info("Rotating file '%s' by %s", filepath, direction)
        direction = Direction.convert_from_string(direction)
        output_file = filepath + '.tmp'

        with open(filepath, 'rb') as f:
            pdf_reader = PyPDF2.PdfFileReader(f)
            pdf_writer = PyPDF2.PdfFileWriter()

            for pagenum in range(pdf_reader.numPages):
                page = pdf_reader.getPage(pagenum)
                page.rotateCounterClockwise(direction.value)
                pdf_writer.addPage(page)

            log.info("Writing PDF: %s", output_file)
            with open(output_file, 'wb') as out:
                pdf_writer.write(out)
        os.remove(filepath)
        shutil.move(output_file, filepath)

    '''
    Converts 'a   b,c , d and e' to 'a','b','c','d','e'
    '''
    def sanitize_names_str(self, names_str):
        tmp_str = names_str.replace(" and ", ",")
        return [tmp_str2.strip() for tmp_str2 in tmp_str.split(',')]

    '''
    Reads excel from loc and returns [headers[] , data[]] 
    '''
    def read_excel_to_memory(self, loc):
        wb_obj = openpyxl.load_workbook(loc, data_only=True)   
        sheet_obj = wb_obj.active

        # Get headers
        excel_headers = []
        excel_rows = []
        for j in range(1, sheet_obj.max_column + 1):
            excel_headers.append(str(sheet_obj.cell(row = 1, column = j).value))
        for i in range(2, sheet_obj.max_row + 1):
            cells = []

            for j in range(1, len(excel_headers) + 1):
                cells.append(str(sheet_obj.cell(row = i, column = j).value))
            
            if all(c == 'None' for c in cells):
                break

            #tmp_str = self.sanitize_names_str(cells[1])
            #tmp_str2 = [i.split()[0] for i in tmp_str]
            #if len(tmp_str2) == 1:
            #    tmp_str = tmp_str2[0]
            #else:
            #    tmp_str3 = ', '.join(tmp_str2[:-1])
            #    tmp_str = tmp_str3 + " and " + tmp_str2[-1]
            #cells.append(tmp_str)
            excel_rows.append(cells)
            
        assert len(excel_rows) > 0,            \
               "**** There are no entries in the excel sheet! ****"
        return [excel_headers, excel_rows]

    def _find_n_files_by_fuzzymatch(self, directory, fuzzystring, filetype = 'pdf', n = 3, ignore_phrases = None):
        candidates = {}

        for filename in os.listdir(directory):
            if filename.endswith(filetype):
                filename_tmp = filename.lower()
                if ignore_phrases:
                    for phrase in ignore_phrases:
                        filename_tmp = filename_tmp.replace(phrase.lower(), '')
                ratio = fuzz.ratio(filename_tmp, fuzzystring.lower())
                candidates[filename] = ratio
        sorted_candidates = sorted(candidates.items(), key=operator.itemgetter(1), reverse=True)
        return [i[0] for i in sorted_candidates][:n]

    def find_n_files_by_fuzzymatch(self, directory, fuzzystring, filetype = 'pdf', n = 3, ignore_phrases = None):
        fuzzystring_set = self.sanitize_names_str(fuzzystring)
        fuzzymatch_res = []
        for f in fuzzystring_set:
            fuzzymatch_res.append(self._find_n_files_by_fuzzymatch(directory, f, filetype, n, ignore_phrases))

        res = []
        for f in fuzzymatch_res:
            res.extend(f)

        return list(set(res))

    '''
    #no in st will be replaced by l[no]
    '''
    def template_to_str(self, st, l):
        log.debug('template_to_str() str = %s' % (st))

        # Find all substitutes to be replaced by a string from the list
        # For e.g. substitutes = ['#1', '#3', '#1']
        substitute = re.findall(r'#\d+', st)
    
        # Remove all duplicates
        # substitutes = ['#1', '#3']
        substitute = list(set(substitute))

        # Get index of each substitute
        # index = [1,3]
        index = [int(s[1:]) for s in substitute]

        # Iterate over each index and replace '#i' with l[i] in the input string
        for i in range(len(index)):
            st = st.replace(substitute[i], l[index[i]])

        log.debug('template_to_str() final str = %s' % st)
        return st

    def save_to_excel(self, excel_data, out_file):
        wb=openpyxl.Workbook()
        ws_write = wb.active

        for row in excel_data:
           ws_write.append(row)

        wb.save(filename=out_file)
        return os.path.join(os.getcwd(), out_file)

def add_next_class_to_excel(input_excel, output_excel):
    f = FileUtils()
    excel_headers, excel_rows = f.read_excel_to_memory(os.path.join("C:\\", "Users", "naira11", "Desktop", "SnehMail", "Clean_sheet.xlsx"))
    for i in range(0, len(excel_rows)):
        if 'Nur' in excel_rows[i][2] or 'nur' in excel_rows[i][2]:
            excel_rows[i].append('Jr.Kg.')
        if 'Jr' in excel_rows[i][2] or 'jr' in excel_rows[i][2] or 'unior' in excel_rows[i][2]:
            excel_rows[i].append('Sr.Kg.')
    data = [excel_headers]
    data.extend(excel_rows)
    return f.save_to_excel(data, output_excel)

if __name__ == "__main__":
    f = FileUtils()
    print(f.find_n_files_by_fuzzymatch(os.path.join("C:\\", "Users", "naira11", "Documents", "wittymail_data", 'allpdf'), 'aarahya jadav,ayush kurund,aarohiBhagat,aahil sayyed', ignore_phrases = ['Nursery', 'Kothrud']))
    #f.pdf_rotate(os.path.join("C:\\", "Users","naira11", 'Test.pdf'), Direction.CLOCKWISE)
    #f.pdf_rotate(Direction.ANTICLOCKWISE)
    #print(add_next_class_to_excel(os.path.join("C:\\", "Users", "naira11", "Desktop", "SnehMail", "Clean_sheet.xlsx"), 'test1.xlsx'))
    next
